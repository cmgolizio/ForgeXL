"""The repaired parser must preserve values through both real Actions and exports."""

import io

import openpyxl
import polars as pl
import pytest

from tests.helpers import xlsx_bytes


@pytest.mark.parametrize("action_id, slot_id, output_id", [
    ("exact_duplicate_remover", "source_file", "deduplicated_data"),
    ("product_master_builder", "sales_file", "product_master"),
])
def test_mixed_values_survive_action_preview_and_both_downloads(
    client, action_id, slot_id, output_id,
):
    header = ["SKU", "Vintage", "Supplier", "Producer", "Selection", "Volume"]
    prefix = ["001", "2025", "Supplier", "Producer", "Selection"]
    rows = [prefix + [value] for value in (10, "n/a", None, "n/a", 2.5)]
    expected = [tuple(prefix + [value]) for value in ("10", "n/a", None, "2.5")]
    response = client.post(
        "/api/runs", data={"action_id": action_id},
        files={slot_id: ("mixed.xlsx", xlsx_bytes({"Data": [header, *rows]}))},
    )
    assert response.status_code == 200, response.text
    manifest = response.json()
    assert manifest["inputs"][0]["parser_engine"] == "openpyxl"
    assert manifest["audit"]["rows_affected"] == 1
    (warning,) = manifest["validation"]["warnings"]
    assert warning["code"] == "MIXED_COLUMN_TYPES"
    assert warning["details"]["columns"] == ["Volume"]
    assert manifest["audit"]["warnings"] == [warning]
    base = f'/api/runs/{manifest["run_id"]}/outputs/{output_id}'
    preview = client.get(f"{base}/preview").json()
    assert [tuple(row) for row in preview["rows"]] == expected
    csv = client.get(f"{base}/download/csv")
    assert csv.status_code == 200
    assert pl.read_csv(csv.content, infer_schema=False).rows() == expected
    xlsx = client.get(f"{base}/download/xlsx")
    assert xlsx.status_code == 200
    book = openpyxl.load_workbook(io.BytesIO(xlsx.content), read_only=True, data_only=True)
    try:
        sheet = book[book.sheetnames[0]]
        assert list(sheet.iter_rows(min_row=2, values_only=True)) == expected
    finally:
        book.close()
