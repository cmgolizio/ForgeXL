"""Generic tabular parsing for every Action (build plan 3.4-3.6 and 6C.6-6C.7).

One entry point, :func:`parse_tabular_bytes`, turns the bytes of an upload into
a Polars dataframe plus the metadata the manifest records: which engine read
the file, which worksheet it came from, and the resulting shape.

Since Phase 6C the input is the uploaded bytes themselves, never a path: an
upload is parsed from memory and is never written out in order to be read back
in. The two Excel engines want different things from those bytes and both are
given what they accept — ``bytes`` for fastexcel, a :class:`io.BytesIO` buffer
for openpyxl — rather than a temporary file standing in for either.

Accuracy rules this module follows:

* **Nothing is guessed silently.** A parse failure is reported as a parse
  failure. The module never retries with progressively stranger delimiters or
  settings in the hope that one sticks (build plan 3.5).
* **No value is silently discarded or normalised.** Mixed Excel cell types
  are preserved as text, with the affected columns reported to the runner
  for a visible warning. No trimming or date repair is applied
  (build plan section 3.3).
* **No engine switch is hidden.** If the preferred Excel engine fails and the
  compatibility fallback succeeds, the fallback is what gets recorded
  (build plan section 6.2).
"""

from __future__ import annotations

import io
from collections.abc import Iterable, Iterator
from dataclasses import dataclass
from typing import Any

import fastexcel
import polars as pl

from app.errors import (
    AmbiguousWorkbookError,
    FileParseError,
    UnsupportedExtensionError,
)

CSV_EXTENSION = ".csv"
XLSX_EXTENSION = ".xlsx"

#: Extensions the POC accepts. Everything else is rejected explicitly rather
#: than attempted hopefully (build plan section 16).
SUPPORTED_EXTENSIONS: tuple[str, ...] = (CSV_EXTENSION, XLSX_EXTENSION)

#: Engine identifiers recorded in the manifest.
ENGINE_POLARS_CSV = "polars-csv"
ENGINE_FASTEXCEL = "fastexcel-calamine"
ENGINE_OPENPYXL = "openpyxl"


class _WorkbookStructureError(FileParseError):
    """A refusal about what the workbook contains, not an engine limitation.

    The compatibility fallback would reach the same conclusion, so these
    propagate immediately instead of triggering a retry.
    """


@dataclass(frozen=True)
class ParsedFile:
    """A parsed upload and how it was read."""

    frame: pl.DataFrame

    #: Which engine actually read the file, including a fallback if one was used.
    parser_engine: str

    #: Worksheet the rows came from; ``None`` for CSV.
    worksheet: str | None = None

    #: Columns preserved as text because their Excel cells have incompatible types.
    mixed_columns: tuple[str, ...] = ()

    @property
    def row_count(self) -> int:
        return self.frame.height

    @property
    def column_count(self) -> int:
        return self.frame.width

    @property
    def columns(self) -> tuple[str, ...]:
        return tuple(self.frame.columns)


def parse_tabular_bytes(payload: bytes, extension: str) -> ParsedFile:
    """Read `payload` as tabular data, dispatching on `extension`.

    Args:
        payload: The uploaded bytes, held in memory (build plan 6C.3).
        extension: Lowercase extension including the leading dot.

    The extension chooses the reader; it does not decide the outcome. A file
    named ``.csv`` that is not a CSV fails here rather than being accepted on
    the strength of its name, which is what build plan 6C.5 means by not
    trusting client-supplied metadata alone.

    Raises:
        UnsupportedExtensionError: the extension is not supported.
        FileParseError: the bytes could not be read as tabular data.
        AmbiguousWorkbookError: the workbook holds more than one data sheet.
    """
    normalised = extension.lower()
    if normalised == CSV_EXTENSION:
        return _parse_csv(payload)
    if normalised == XLSX_EXTENSION:
        return _parse_xlsx(payload)
    raise UnsupportedExtensionError(
        f"{normalised or 'That file type'} is not a supported file type. "
        f"Supported types are {_human_list(SUPPORTED_EXTENSIONS)}.",
        details={
            "extension": normalised,
            "supported_extensions": list(SUPPORTED_EXTENSIONS),
        },
    )


# ---------------------------------------------------------------------------
# CSV (build plan 3.5)
# ---------------------------------------------------------------------------


def _parse_csv(payload: bytes) -> ParsedFile:
    """Read CSV bytes with Polars.

    Polars reads the buffer directly (build plan 6C.6). Its own delimiter and
    type inference is used as-is. Date parsing is deliberately left off: a
    column of date-shaped text stays text unless an Action asks for something
    else, so no value is silently retyped.
    """
    try:
        frame = pl.read_csv(payload)
    except pl.exceptions.NoDataError as exc:
        raise FileParseError(
            "The uploaded CSV file is empty.", details={"reason": str(exc)}
        ) from exc
    except Exception as exc:
        raise FileParseError(
            "The uploaded CSV file could not be read. It may be malformed or "
            "may not be a CSV file.",
            details={"reason": str(exc)},
        ) from exc
    return ParsedFile(frame=frame, parser_engine=ENGINE_POLARS_CSV)


# ---------------------------------------------------------------------------
# XLSX (build plan 3.6 and section 17)
# ---------------------------------------------------------------------------


def _parse_xlsx(payload: bytes) -> ParsedFile:
    """Read a single-data-sheet XLSX workbook from memory.

    fastexcel (calamine) is the preferred engine. If it cannot open the
    workbook accurately, openpyxl is tried as the compatibility fallback and
    whichever engine succeeded is what the manifest records.

    A refusal about the workbook's *structure* — no data sheet, or several —
    is never retried: the fallback would reach the same conclusion.

    Both engines read the values stored in the workbook. Neither evaluates
    formulas and neither executes macros (build plan section 17 and 7F).
    """
    try:
        return _parse_xlsx_with_fastexcel(payload)
    except (AmbiguousWorkbookError, _WorkbookStructureError):
        raise
    except Exception as primary_error:
        try:
            return _parse_xlsx_with_openpyxl(payload)
        except (AmbiguousWorkbookError, _WorkbookStructureError):
            raise
        except Exception as fallback_error:
            raise FileParseError(
                "The uploaded Excel workbook could not be read. It may be "
                "corrupt, password-protected, or not a genuine .xlsx file.",
                details={
                    "primary_engine": ENGINE_FASTEXCEL,
                    "primary_reason": str(primary_error),
                    "fallback_engine": ENGINE_OPENPYXL,
                    "fallback_reason": str(fallback_error),
                },
            ) from fallback_error


def _parse_xlsx_with_fastexcel(payload: bytes) -> ParsedFile:
    """Read the workbook with the preferred engine, straight from the bytes.

    fastexcel accepts ``bytes`` but rejects a file-like buffer, so the payload
    is handed over unwrapped (build plan 6C.7).

    Every worksheet is probed before one is chosen, and a probe that fails is
    an engine failure like any other: it propagates out of this function and
    into :func:`_parse_xlsx`, which retries the whole workbook with the
    compatibility fallback. A worksheet this engine cannot open is never
    counted as an empty one — see :func:`_fastexcel_sheet_has_data`.
    """
    reader = fastexcel.read_excel(payload)
    worksheet = _select_data_worksheet(
        {name: _fastexcel_sheet_has_data(reader, name) for name in reader.sheet_names}
    )
    sheet = reader.load_sheet(
        worksheet, header_row=0, schema_sample_rows=None, dtype_coercion="strict"
    )
    frame = sheet.to_polars()
    # Even strict coercion in the pinned reader can treat literal text such as
    # "n/a" as null in a numeric column. Compare null positions with an explicit
    # text read of just the nullable columns. Real blanks stay null in both;
    # any newly blanked value makes this an engine failure, invoking the
    # existing, recorded openpyxl fallback. No all-string frame reaches an
    # Action through this check, and non-nullable sheets need no second read.
    nullable = [column.name for column in frame if column.null_count()]
    if nullable:
        source_text = reader.load_sheet(
            worksheet, header_row=0, use_columns=nullable, dtypes="string"
        ).to_polars()
        if source_text.height != frame.height or any(
            (frame[name].is_null() & source_text[name].is_not_null()).any()
            for name in nullable
        ):
            raise ValueError("The preferred XLSX reader would discard cell values.")
    return ParsedFile(
        frame=frame,
        parser_engine=ENGINE_FASTEXCEL,
        worksheet=worksheet,
    )


def _fastexcel_sheet_has_data(reader: fastexcel.ExcelReader, name: str) -> bool:
    """Whether a worksheet contains any cells at all.

    ``header_row=None`` counts every row, so a sheet holding only a header row
    still counts as a data sheet — a header-only upload is a legitimate dataset
    with zero rows, not an empty sheet.

    A worksheet that genuinely holds nothing reports a height and width of
    zero; it does not fail to load. So a load failure here is not evidence
    about the sheet's contents, and reporting one as "empty" would be a guess —
    a guess with two silent consequences: a workbook whose only data sheet
    failed to open would be refused as containing no data, and a workbook with
    one readable sheet beside it would have that other sheet selected as if it
    were unambiguous. Neither is true, and neither is visible to the user.

    So nothing is caught here. The exception propagates into the engine
    fallback in :func:`_parse_xlsx`, which reads the workbook again with
    openpyxl and, if that fails too, reports a parse error naming both engines
    (build plan 3.5 and section 6.2).
    """
    sheet = reader.load_sheet(name, header_row=None)
    return sheet.total_height > 0 and sheet.width > 0


def _parse_xlsx_with_openpyxl(payload: bytes) -> ParsedFile:
    """Read the workbook with the compatibility fallback engine.

    openpyxl wants a file-like object rather than raw bytes, so the payload is
    wrapped in an in-memory buffer — not written to a temporary file
    (build plan 6C.7).

    ``data_only=True`` returns the values Excel last stored for formula cells
    rather than the formula text. openpyxl neither evaluates formulas nor runs
    macros.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(
        io.BytesIO(payload), read_only=True, data_only=True, keep_links=False
    )
    mixed_columns: list[str] = []
    try:
        worksheet_name = _select_data_worksheet(
            {
                name: _openpyxl_sheet_has_data(workbook[name])
                for name in workbook.sheetnames
            }
        )
        frame = _frame_from_rows(
            workbook[worksheet_name].iter_rows(values_only=True),
            mixed_columns=mixed_columns,
        )
    finally:
        workbook.close()

    return ParsedFile(
        frame=frame,
        parser_engine=ENGINE_OPENPYXL,
        worksheet=worksheet_name,
        mixed_columns=tuple(mixed_columns),
    )


def _openpyxl_sheet_has_data(worksheet: Any) -> bool:
    """Whether an openpyxl worksheet holds any cells.

    In read-only mode the recorded dimensions come from the workbook's own
    dimension tag, which some writers omit. When they are unavailable the sheet
    is probed by reading its first row rather than assumed empty.
    """
    max_row = getattr(worksheet, "max_row", None)
    max_column = getattr(worksheet, "max_column", None)
    if max_row is not None and max_column is not None:
        return max_row > 0 and max_column > 0
    return any(True for _ in worksheet.iter_rows(max_row=1, values_only=True))


def _frame_from_rows(
    rows: Iterable[tuple[Any, ...]], *, mixed_columns: list[str] | None = None
) -> pl.DataFrame:
    """Build a dataframe from openpyxl's row tuples, first row as the header."""
    iterator: Iterator[tuple[Any, ...]] = iter(rows)
    try:
        header = next(iterator)
    except StopIteration:
        raise _WorkbookStructureError(
            "The selected worksheet is empty.", details={"worksheet_empty": True}
        ) from None

    column_names = [
        str(value) if value is not None else f"column_{index + 1}"
        for index, value in enumerate(header)
    ]

    data_rows: list[list[Any]] = []
    for row in iterator:
        values = list(row[: len(column_names)])
        # Pad short rows so every record matches the header width.
        values.extend([None] * (len(column_names) - len(values)))
        data_rows.append(values)

    # One Polars column has one type. Preserve incompatible Excel cell types
    # as text instead of letting construction coerce booleans into numbers or
    # discard strings. Integers and floats remain one numeric family.
    for index in range(len(column_names)):
        kinds = {
            "number" if type(row[index]) in (int, float) else type(row[index])
            for row in data_rows
            if row[index] is not None
        }
        if len(kinds) > 1:
            if mixed_columns is not None:
                mixed_columns.append(column_names[index])
            for row in data_rows:
                value = row[index]
                if value is not None:
                    row[index] = str(value)

    return pl.DataFrame(
        data_rows or None,
        schema=column_names,
        orient="row",
        infer_schema_length=None,
    )


def _select_data_worksheet(sheets: dict[str, bool]) -> str:
    """Return the one worksheet holding data, or refuse to choose.

    Build plan section 17: the POC must not silently pick a worksheet from a
    workbook containing several plausible data sheets. Exactly one data sheet
    is unambiguous; anything else is reported to the user.

    Raises:
        _WorkbookStructureError: no worksheet contains data.
        AmbiguousWorkbookError: more than one worksheet contains data.
    """
    populated = [name for name, has_data in sheets.items() if has_data]

    if len(populated) == 1:
        return populated[0]

    if not populated:
        raise _WorkbookStructureError(
            "The uploaded Excel workbook contains no data.",
            details={"worksheets": list(sheets)},
        )

    raise AmbiguousWorkbookError(
        "This workbook contains multiple worksheets. The POC currently "
        "requires a workbook containing one data worksheet. Save the required "
        "worksheet as its own workbook or CSV and try again.",
        details={"worksheets_with_data": populated, "worksheets": list(sheets)},
    )


def _human_list(values: tuple[str, ...]) -> str:
    """Join `values` for a user-facing message: 'a, b and c'."""
    if len(values) <= 1:
        return "".join(values)
    return f"{', '.join(values[:-1])} and {values[-1]}"